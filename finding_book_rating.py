# -*- coding: utf-8 -*-
"""
Created on Fri Jan 10 18:08:36 2020

@author: pratms
"""
import time
start = time.time()
import bs4,requests
import openpyxl
try: 
    from googlesearch import search 
except ImportError:  
    print("No module named 'google' found")

path=r"C:\Users\pratms\pythonprojects\webscrapping\webscrapping_medium_top_10\excel_doc1.xlsx"

wb_obj=openpyxl.load_workbook(path)
sheet_obj=wb_obj.active
m_row=sheet_obj.max_row
print("rows="+str(m_row))
'''
for i in range(1,m_row+1):
    bookname=sheet_obj.cell(row=i,column=2).value
    bookauthor=sheet_obj.cell(row=i,column=3).value
    #searchPhrase=bookname+" by "+bookauthor+" in good reads"
    searchPhrase="Catch-22 by Joseph Heller in good reads"
    print(searchPhrase)
'''
     
searchPhrase="Catch-22 by Joseph Heller in good reads"  
for j in search(searchPhrase, tld="com", num=10, stop=1, pause=3): 
    goodreadsurl=j
    print(goodreadsurl)


    
end = time.time()
print("time taken by program is:"+str(end - start)) 