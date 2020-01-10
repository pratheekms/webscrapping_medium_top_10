# -*- coding: utf-8 -*-
"""
Created on Fri Jan 10 19:05:37 2020

@author: pratms
"""
import bs4,requests
#import lxml
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
try: 
    from googlesearch import search 
except ImportError:  
    print("No module named 'google' found") 
  
# to search 
import openpyxl

path=r"C:\Users\pratms\pythonprojects\webscrapping\webscrapping_medium_top_10\excel_doc1.xlsx"

wb_obj=openpyxl.load_workbook(path)
sheet_obj=wb_obj.active
m_row=sheet_obj.max_row
print("rows="+str(m_row))
for i in range(1,m_row+1):
    bookname=sheet_obj.cell(row=i,column=2).value
    bookauthor=sheet_obj.cell(row=i,column=3).value
    searchPhrase=bookname+" by "+bookauthor+" goodreads"
    #searchPhrase="Catch-22 by Joseph Heller in good reads"
    print(searchPhrase)
#searchPhrase="1984 by George Orwell in goodreads" 
    print(searchPhrase) 
    for j in search(searchPhrase, tld="com", num=10, stop=1, pause=3): 
        goodreadsurl=j
    print(goodreadsurl)
    print("bs4 starting")
    res=requests.get(goodreadsurl,verify=False)
    res.raise_for_status()
    soup=bs4.BeautifulSoup(res.text,'lxml')
#rating=soup.find_all("span",attrs={"itemprop":"ratingValue"})
    ratingObj=soup.find("span", itemprop="ratingValue")
    rating=ratingObj.getText()
    print(rating)
    sheet_obj.cell(row=i,column=4).value=str(rating)
    print("rating write complete")
    wb_obj.save(r"C:\Users\pratms\pythonprojects\webscrapping\webscrapping_medium_top_10\excel_doc1.xlsx")