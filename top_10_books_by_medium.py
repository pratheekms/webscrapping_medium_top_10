# -*- coding: utf-8 -*-
"""
Created on Thu Jan  9 16:38:56 2020

@author: pratheek m s
"""
import time
start = time.time()
import bs4,requests
import openpyxl
import random
import logging

from datetime import datetime

GlobalPath=r"C:\Users\prath\pythonproj\webscrapping_medium_top_10"
LOG_FILENAME = datetime.now().strftime('logfile_%H_%M_%S_%d_%m_%Y.log')
logging.basicConfig(filename=LOG_FILENAME, 
                    format='%(asctime)s %(message)s', 
                    filemode='w')

#Creating an object 
logger=logging.getLogger() 
  
#Setting the threshold of logger to DEBUG 
logger.setLevel(logging.DEBUG) 
'''  
#Test messages 
logger.debug("Harmless debug Message") 
logger.info("Just an information") 
logger.warning("Its a Warning") 
logger.error("Did you try to divide by zero") 
logger.critical("Internet is down")
'''


#path=GlobalPath+r"\excel_doc1.xlsx"
#open excel file code start
print("Global path is",GlobalPath)
logger.info("Global path is",GlobalPath)
print("opening new excel for book author rating")
logger.info("opening new excel for book author rating")
wb_objex = openpyxl.Workbook()
sheet_objex = wb_objex.active
sheet_objex.cell(row=1,column=1).value='Sl. No'
sheet_objex.cell(row=1,column=2).value='Book Name'
sheet_objex.cell(row=1,column=3).value='Book Author'
sheet_objex.cell(row=1,column=4).value='Rating'

print("writing of headers to the new excel sheet")
logger.info("writing of headers to the new excel sheet")

finalResult=[]
bookdict={}
bookcount=1
print("scraping web for top 100 books start") 
logger.info("------scraping web for top 100 books start------")   
url="https://medium.com/world-literature/creating-the-ultimate-list-100-books-to-read-before-you-die-45f1b722b2e5"
res=requests.get(url,verify=False)
res.raise_for_status()
soup=bs4.BeautifulSoup(res.text,'lxml')
#slno=soup.find_all("a",attrs={"class":"fi cn hx hy hz ia"})
bookAndAuthors=soup.find_all("strong",attrs={"class":"id ke"})


print("------scraping web for top 100 books end------")
logger.info("------scraping web for top 100 books end------")

print("exctrating data into dict start")
logger.info("exctrating data into dict start")

for i in range(0,len(bookAndAuthors)-4):
    if len(bookAndAuthors[i].getText())>2:
        finalResult.append(bookAndAuthors[i].getText())
        print(bookAndAuthors[i].getText())
        

for i in finalResult:
    print("book "+str(bookcount)+i)
    logger.info("book "+str(bookcount)+i)
    print(len(finalResult[bookcount-1]))
   
    bookdict.update({bookcount:{'name':i.split(' by ')[0],'author':i.split(' by ')[1]}})
    bookcount+=1
print("exctrating data into dict end")
logger.info("exctrating data into dict end")
print("exctrating from dict to list")
logger.info("exctrating from dict to list")
print("------writng list slno book author to excel start------")
logger.info("------writng list slno book author to excel start------")
for bn,ba in bookdict.items():
    print("book num",bn)
    
    booknum=bn
    for key in ba:
        print(key+":",ba[key])
        
        bookname=ba['name']
        bookauthor=ba['author']
    f = open(GlobalPath+r"\top100BokksByMedium.txt", "a")
    f.write(str(booknum)+'\t'+str(bookname)+'\t'+bookauthor+'\n')
    f.close()
    sheet_objex.cell(row=booknum+1,column=1).value=str(booknum)
    sheet_objex.cell(row=booknum+1,column=2).value=str(bookname)
    sheet_objex.cell(row=booknum+1,column=3).value=str(bookauthor)
    #sheet_objex.cell(row=booknum+1,column=4).value='Rating'
print("------writng list slno book author to excel end------")
logger.info("------writng list slno book author to excel end------")
wb_objex.save(GlobalPath+r"\top100BokksByMedium.xlsx")
print("------excel with with slno book author saved------")
logger.info("------excel with with slno book author saved------")
print("------------rating process starts------------")
logger.info("------------rating process starts------------")
#book name and book author details are saved to an excel file
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
try: 
    from googlesearch import search 
except ImportError:  
    print("No module named 'google' found") 
    logger.info("Just an information")


pathForRating=GlobalPath+r"\top100BokksByMedium.xlsx"
print("------saved excel file open------")
logger.info("------saved excel file open------")
wb_obj=openpyxl.load_workbook(pathForRating)
sheet_obj=wb_obj.active
m_row=sheet_obj.max_row
print("rows="+str(m_row))
logger.info("------total rows in excel file="+str(m_row),"------")
logger.info("---------------------------------------------------")
for i in range(1,m_row):
    logger.info("---------------------------------------------------")
    bookname=sheet_obj.cell(row=i+1,column=2).value
    bookauthor=sheet_obj.cell(row=i+1,column=3).value
    searchPhrase=bookname+" by "+bookauthor+" goodreads"
    #searchPhrase="Catch-22 by Joseph Heller in good reads"
    print("modified search phrase ",searchPhrase)
    logger.info("book num=",i,'\t',bookname,'\t',bookauthor)
    logger.info("modified search phrase --->",searchPhrase)
#searchPhrase="1984 by George Orwell in goodreads" 
    #print(searchPhrase) 
    for j in search(searchPhrase, tld="com", num=10, stop=1, pause=random.randint(1,20)): 
        goodreadsurl=j
    print("good reads URL for ",bookname,bookauthor,"is ",goodreadsurl)
    logger.info("good reads URL for ",bookname,bookauthor,"is --->",goodreadsurl)
    
    print("---scrapping for rating start---")
    logger.info("---scrapping for rating start---")
    res=requests.get(goodreadsurl,verify=False)
    res.raise_for_status()
    soup=bs4.BeautifulSoup(res.text,'lxml')
    ratingObj=soup.find("span", itemprop="ratingValue")
    try:
        rating=ratingObj.getText()
        print("ratinng of",bookname,bookauthor, rating)
        logger.info("ratinng of",bookname,bookauthor, rating)
        sheet_obj.cell(row=i+1,column=4).value=str(rating)
        print("---rating write complete---")
        logger.info("---rating write complete---")
        
    except:
        print("ratinng of",bookname,bookauthor,"Error")
        logger.info("ratinng of",bookname,bookauthor,"Error")
        sheet_obj.cell(row=i+1,column=4).value="Error"
        print("rating write complete with ERROR")
        logger.info("^^^rating write complete with ERROR^^^")
        print("-----------------------")
        wb_obj.save(pathForRating)
        print("final excel saved")
        logger.info("---final excel saved---")
        logger.info("---------------------------------------------------")
  
#wb_objex.save(r"C:\Users\pratms\pythonprojects\webscrapping\webscrapping_medium_top_10\top100BokksByMedium.xlsx")
   
end = time.time()
print("time taken by program is:"+str(end - start)) 
logger.info("____time taken by program is:"+str(end - start)+"____")
logging.shutdown()