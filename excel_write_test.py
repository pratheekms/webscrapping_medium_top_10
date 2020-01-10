# -*- coding: utf-8 -*-
"""
Created on Fri Jan 10 21:03:34 2020

@author: pratms
"""

import openpyxl

import logging 
import time


GlobalPath=r"C:\Users\pratms\pythonprojects\webscrapping\webscrapping_medium_top_10"
path=GlobalPath+r"\excel_doc1.xlsx"
newfile=GlobalPath+'\\'+str(time.ctime(time.time()))+'.log'
logging.basicConfig(filename=newfile, 
                    format='%(asctime)s %(message)s', 
                    filemode='w')

#Creating an object 
logger=logging.getLogger() 
  
#Setting the threshold of logger to DEBUG 
logger.setLevel(logging.DEBUG) 
  
#Test messages 
logger.debug("Harmless debug Message") 
logger.info("Just an information") 
logger.warning("Its a Warning") 
logger.error("Did you try to divide by zero") 
logger.critical("Internet is down")

wb_obj=openpyxl.load_workbook(path)
sheet_obj=wb_obj.active
m_row=sheet_obj.max_row
print("rows="+str(m_row))
for i in range(1,m_row+1):
    bookname=sheet_obj.cell(row=i,column=2).value
    bookauthor=sheet_obj.cell(row=i,column=3).value
    searchPhrase=bookname+" by "+bookauthor+" goodreads"
    #searchPhrase="Catch-22 by Joseph Heller in good reads"
    print(searchPhrase)
    logger.info("Just an information",searchPhrase)
    rating=10
    sheet_obj.cell(row=i,column=4).value=str(rating)
    print("rating write complete")
    logger.info("rating write complete")

wb_obj.save(path)
logging.shutdown()